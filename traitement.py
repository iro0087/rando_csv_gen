from openpyxl import load_workbook

from pyexcel.cookbook import merge_all_to_a_book

import glob, os

csv_file = str(input("What is the file?"))

output_file = str(input("What is the name of the output file?"))

step = float(input("What is the step?"))

min_ = float(input("What is min value?"))

xl_file = csv_file + ".xlsx"

merge_all_to_a_book(glob.glob(csv_file), xl_file)

w = load_workbook(xl_file)

sheet = w.active

cc = int(input("What is the column ?"))

listl = []

listl2 = []

listl2b = []

listl3 = []

t = 1

while sheet.cell(row=t, column=cc).value != None:

    listl.append(float(sheet.cell(row=t, column=cc).value))

    t += 1 

t = 1

while min(listl) != max(listl):

    listl2.append(min(listl))

    listl[listl.index(min(listl))] = max(listl)

listl2.append(listl[0])

listl2b.append(listl2[0])

listl3.append(listl2.count(listl2[0]))

while t < len(listl2):

    if listl2[t] != listl2[t - 1]:

        listl2b.append(listl2[t])

        listl3.append(listl2.count(listl2[t]))

    t += 1

t = 0

t2 = 1

w = load_workbook(output_file)

sheet = w.active

while t < len(listl3):

    sheet.cell(row=t+1, column=1).value = min_

    min_ += step

    t += 1

t = 1

listF = [listl2b, listl3]

while t < 3:

    while t2 < len(listl3) + 1:

        sheet.cell(row=t2, column=t + 1).value = listF[t - 1][t2 - 1]

        t2 += 1

    t2 = 1

    t += 1

w.save(output_file)

print("done :)")



